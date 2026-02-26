#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, List, Dict, Any, Optional

from openpyxl import Workbook


# ========= Leitura do PDF (pdfplumber -> fallback pypdf) =========
def extract_pages_text(pdf_path: Path) -> Iterator[str]:
    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                yield page.extract_text() or ""
        return
    except Exception:
        pass

    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(pdf_path))
        for page in reader.pages:
            yield page.extract_text() or ""
    except Exception as exc:
        raise RuntimeError("Não foi possível ler o PDF. Instale 'pdfplumber' ou 'pypdf'.") from exc


def norm(s: str) -> str:
    s = (s or "").replace("\u00a0", " ").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def parse_number_ptbr(s: str) -> Optional[float]:
    s = (s or "").strip()
    if not s:
        return None
    s = re.sub(r"[^0-9.,-]", "", s)
    if not s:
        return None
    # se tiver vírgula, assume vírgula como decimal
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def fmt_int_or_float(v: Optional[float]) -> str:
    if v is None:
        return ""
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    out = f"{v:.4f}".rstrip("0").rstrip(".")
    return out


# ========= descrição robusta =========
def extract_descricao(block: str) -> str:
    """
    Preferência: pega o TÍTULO do item (logo abaixo de 'DETALHAMENTO DO ITEM 000xx'),
    que é a descrição curta (melhor para tabela).
    Fallback: pega 'Descrição detalhada' até o próximo marcador, sem engolir o resto do PDF.
    """
    b = (block or "").replace("\u00a0", " ").replace("\r", "\n")
    b = re.sub(r"[ \t]+", " ", b)

    # ---------- 1) TÍTULO (curto) ----------
    m_item = re.search(r"DETALHAMENTO DO ITEM\s+\d+\s*\n", b, flags=re.IGNORECASE)
    if m_item:
        tail = b[m_item.end():]

        m_stop = re.search(
            r"(?im)^\s*(Descri[cç][aã]o|C[oó]digo do|Tipo do item|Quantidade homologada|Vig[êe]ncia inicial|FORNECEDOR\(ES\)|UNIDADE\(S\))",
            tail,
        )
        title_part = tail[: m_stop.start()] if m_stop else tail[:300]
        title_part = re.sub(r"\s+", " ", title_part).strip()

        if title_part and len(title_part) > 10:
            return title_part

    # ---------- 2) FALLBACK: DESCRIÇÃO DETALHADA ----------
    m_desc = re.search(r"(?im)^\s*Descri[cç][aã]o\b", b)
    if not m_desc:
        return ""

    start = m_desc.end()
    tail = b[start:]

    tail = re.sub(r"(?i)\bdetalhada\s*:\s*", " ", tail)

    m_end = re.search(
        r"(?im)^\s*(C[oó]digo\s+do\s*[\n ]*item\s*:?"
        r"|Tipo\s+do\s+item\s*:?"
        r"|Quantidade\s+homologada\s*:?"
        r"|Vig[êe]ncia\s+inicial\s*:?"
        r"|FORNECEDOR\(ES\)"
        r"|UNIDADE\(S\))",
        tail,
    )
    desc = tail[: m_end.start()] if m_end else tail[:400]

    desc = re.sub(r"\s+", " ", desc).strip()
    return desc


# ==================== Detectores / Regex ====================
@dataclass
class AtaMeta:
    numero_ata: str = ""
    vigencia: str = ""
    unidade_gerenciadora: str = ""


RE_UG = re.compile(r"Unidade\s+Gerenciadora\s+([0-9]{6}\s*-\s*[^\n]+)", re.IGNORECASE)
RE_ATA_NUM = re.compile(r"\bn[ºo]\s*([0-9]{1,6}/[0-9]{4})\b", re.IGNORECASE)
RE_VIG = re.compile(
    r"Vig[êe]ncia.*?de\s*([0-3]?\d/[01]?\d/\d{4})\s*a\s*([0-3]?\d/[01]?\d/\d{4})",
    re.IGNORECASE | re.DOTALL,
)

RE_ITEM_START = re.compile(r"DETALHAMENTO\s+DO\s+ITEM\s+([0-9]{1,5})", re.IGNORECASE)

# CRO/2 linha (sempre está no quadro UNIDADE(S))
RE_CRO2 = re.compile(
    r"\b160491\s+CRO/2\s+Participante\s+([0-9][0-9.,]+)\s+([0-9][0-9.,]+)",
    re.IGNORECASE,
)

# Pega a linha do fornecedor e captura o ÚLTIMO número como valor unitário
RE_SUPPLIER_LINE = re.compile(
    r"^\s*\d{3}\s+\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\s+.*?\s+([0-9][0-9.,]+)\s*$",
    re.IGNORECASE | re.MULTILINE,
)


# ==================== Extração de Meta por Ata ====================
def extract_meta_from_page(text: str) -> AtaMeta:
    ug = ""
    ata = ""
    vig = ""

    m = RE_UG.search(text)
    if m:
        ug = norm(m.group(1))

    m = RE_ATA_NUM.search(text)
    if m:
        ata = norm(m.group(1))

    m = RE_VIG.search(text)
    if m:
        vig = f"{m.group(1)} a {m.group(2)}"

    return AtaMeta(numero_ata=ata, vigencia=vig, unidade_gerenciadora=ug)


def is_ata_header_page(text: str) -> bool:
    t = text.upper()
    return ("RELATÓRIO ATA DE REGISTRO DE PREÇOS" in t) and ("INFORMAÇÕES DA ATA" in t)


# ==================== Blocos por Item (mantém páginas seguintes) ====================
def iter_item_blocks(ata_pages: List[str]) -> Iterator[str]:
    full = "\n\n".join(ata_pages)
    parts = re.split(r"(?=DETALHAMENTO DO ITEM\s+\d+)", full, flags=re.IGNORECASE)
    for p in parts:
        p = p.strip()
        if p.upper().startswith("DETALHAMENTO DO ITEM"):
            yield p


def extract_item_data(block: str) -> Optional[Dict[str, Any]]:
    m_item = RE_ITEM_START.search(block)
    if not m_item:
        return None
    item = m_item.group(1).zfill(5)

    desc = extract_descricao(block)

    # valor unitário (pega linha do fornecedor, último número)
    valor_unit = None
    m_sup = RE_SUPPLIER_LINE.search(block)
    if m_sup:
        valor_unit = parse_number_ptbr(m_sup.group(1))

    # CRO/2 quantidades
    m_cro = RE_CRO2.search(block)
    if not m_cro:
        return None

    qtd_reg = parse_number_ptbr(m_cro.group(1))
    qtd_disp = parse_number_ptbr(m_cro.group(2))

    if not qtd_reg or qtd_reg <= 0:
        return None

    return {
        "Item": item,
        "DescricaoItem": desc,
        "ValorUnitario": fmt_int_or_float(valor_unit),
        "QtdRegistrada": fmt_int_or_float(qtd_reg),
        "QtdDisponivelRemanejamento": fmt_int_or_float(qtd_disp),
    }


# ==================== Processamento do PDF inteiro (várias atas) ====================
def process_pdf(pdf_path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    meta = AtaMeta()
    ata_pages: List[str] = []

    def flush_current_ata():
        nonlocal ata_pages, meta, rows
        if not ata_pages:
            return
        for block in iter_item_blocks(ata_pages):
            item_data = extract_item_data(block)
            if not item_data:
                continue
            rows.append(
                {
                    "NumeroAta": meta.numero_ata,
                    "Vigencia": meta.vigencia,
                    "UnidadeGerenciadora": meta.unidade_gerenciadora,
                    **item_data,
                }
            )

    for page_text_raw in extract_pages_text(pdf_path):
        page_text = norm(page_text_raw)
        if not page_text:
            continue

        if is_ata_header_page(page_text):
            flush_current_ata()
            ata_pages = []
            meta = extract_meta_from_page(page_text)

        ata_pages.append(page_text)

    flush_current_ata()
    return rows


# ==================== Exportação XLSX ====================
def save_xlsx(rows: List[Dict[str, Any]], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "CRO2_Itens"

    # ✅ NOVA COLUNA: SaldoRemanejamento
    headers = [
        "NumeroAta",
        "Vigencia",
        "UnidadeGerenciadora",
        "Item",
        "DescricaoItem",
        "ValorUnitario",
        "QtdRegistrada",
        "QtdDisponivelRemanejamento",
        "SaldoRemanejamento",
    ]
    ws.append(headers)

    # escreve linhas + calcula saldo
    for r in rows:
        vu = parse_number_ptbr(str(r.get("ValorUnitario", "") or "")) or 0.0
        qd = parse_number_ptbr(str(r.get("QtdDisponivelRemanejamento", "") or "")) or 0.0
        saldo = vu * qd
        r["SaldoRemanejamento"] = saldo  # guarda também no dict (se precisar)

        ws.append(
            [
                r.get("NumeroAta", ""),
                r.get("Vigencia", ""),
                r.get("UnidadeGerenciadora", ""),
                r.get("Item", ""),
                r.get("DescricaoItem", ""),
                r.get("ValorUnitario", ""),
                r.get("QtdRegistrada", ""),
                r.get("QtdDisponivelRemanejamento", ""),
                round(saldo, 2),  # 2 casas
            ]
        )

    # Ajustes de alinhamento
    wrap = Alignment(wrap_text=True, vertical="center")
    center_v = Alignment(vertical="center")

    max_row = ws.max_row
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=5).alignment = wrap  # DescricaoItem (E)
        for col in (1, 2, 3, 4, 6, 7, 8, 9):
            ws.cell(row=row, column=col).alignment = center_v

        # Formato moeda/2 decimais no saldo (I)
        ws.cell(row=row, column=9).number_format = "0.00"

    # ✅ Mesclar A, B, C quando for a mesma ATA (blocos consecutivos)
    def key_at(row_idx: int):
        return (
            ws.cell(row=row_idx, column=1).value,
            ws.cell(row=row_idx, column=2).value,
            ws.cell(row=row_idx, column=3).value,
        )

    start = 2
    while start <= max_row:
        k = key_at(start)
        end = start
        while end + 1 <= max_row and key_at(end + 1) == k:
            end += 1

        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
            ws.merge_cells(start_row=start, start_column=3, end_row=end, end_column=3)

            for col in (1, 2, 3):
                ws.cell(row=start, column=col).alignment = Alignment(
                    vertical="center", wrap_text=True
                )

        start = end + 1

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrai itens onde CRO/2 é participante (com quantitativo) de PDFs de Atas do Contratos.gov.br."
    )
    parser.add_argument("pdf", type=Path, help="PDF (pode conter várias Atas)")
    parser.add_argument("-o", "--output", type=Path, default=Path("Relatorio.xlsx"), help="XLSX de saída")
    args = parser.parse_args()

    if not args.pdf.exists():
        raise SystemExit(f"Arquivo não encontrado: {args.pdf}")

    rows = process_pdf(args.pdf)
    save_xlsx(rows, args.output)
    print(f" Concluído. Linhas: {len(rows)} | Saída: {args.output.resolve()}")


if __name__ == "__main__":
    main()